import random
from database.connect_db import connect_db
from bson.objectid import ObjectId
from EnToBnValue import EnToBnValue
from openpyxl import Workbook, load_workbook


class InspectionExcel:
    def __init__(self, noteId):
        self.db = connect_db()
        self.note = self.db.notes.find_one({"_id": ObjectId(noteId)})
        self.wb=load_workbook("report.xlsx")

    def company_details(self):
        company = self.db.companies.find_one({"_id": self.note["companyId"]})
        title = company["title"]
        address = company["address"]
        return {"title": title, "address": address}

    def generate_excel_files(self):
        productList = self.note["productList"]
        for product in productList:
            for i in range(len(product["sizes"])):
                productDetails = {
                    "name": product["name"],
                    "brand": product["brand"],
                    "varient": product["varient"],
                    "unit": product["unit"],
                    "size": product["sizes"][i],
                    "packetSize": product["packetSizes"][i]
                }
                self.generate_ins_report(productDetails)
        del self.wb["template"]
        self.wb.save('newR.xlsx')
    
    def randomized_list(self,size,start_mod=1,end_mod=1):
        sizes_list=[]
        for i in range(32):
            sizes_list.append(round(random.uniform(size*start_mod,size*end_mod), 2))
        sizes_in_bn = []

        for item in sizes_list:
            val = EnToBnValue(item).in_bn_letter()
            sizes_in_bn.append(val)

        return sizes_in_bn

    def generate_ins_report(self, productDetails):
        name, brand, varient, unit, size, packetSize = [productDetails[k] for k in (
            'name', 'brand', 'varient', 'unit', 'size', 'packetSize')]
        company_details = self.company_details()
        inspection_date = self.note["inspectionDate"]
        
        def calc_allowable_limit(size):
            if size<=50:
                return size*0.09
            elif size<=100:
                return 4.5
            elif size<=200:
                return size*0.045
            elif size<=300:
                return 9
            elif size<=500:
                return size*0.03
            elif size<=1000:
                return 15
            elif size<=10000:
                return size*0.015
            elif size<=15000:
                return 150
            else:
                return size*0.01
        
        size = int(size)
        allowable_limit = calc_allowable_limit(size=size)
        ws = self.wb.copy_worksheet(self.wb["template"])
        ws.title=str(size)+" "+unit

        if unit=="গ্রাম":
            ws['C12']= f'গ্রোস ওজন \n{unit}'
            ws['D12']= f'খালি মোড়কের ওজন \n{unit}'
            ws['E12']= f'প্রকৃত ওজন \n{unit}'
        
            packetSize = int(packetSize)
            grossSize = size+packetSize

            gross_in_bn = self.randomized_list(size=grossSize,end_mod=1.03)
            empty_pack_in_bn = self.randomized_list(size=packetSize,start_mod=0.97)
            for i in range(13, 45):
                ws[f'C{i}'].value = float(gross_in_bn[i-13])
            for i in range(13, 45):
                ws[f'D{i}'].value = float(empty_pack_in_bn[i-13])
            for i in range(13, 45):
                ws[f'E{i}'].value = f'=SUM(C{i}-D{i})'
                print(ws[f'E{i}'].value)
            for i in range(13, 45):
                ws[f'G{i}'].value = f'=SUM(E{i}-{size})'
            ws["C45"] = "=AVERAGE(C13:C44)"
            ws["D45"] = "=AVERAGE(D13:D44)"
            ws["E45"] = "=SUM(C45-D45)"
            ws["G45"] = f'=SUM(E45-{size})'
        elif unit == "মিঃ লিঃ":
            ws['E12']= f'প্রকৃত পরিমান \n{unit}'
            volumes_in_bn=self.randomized_list(size=size)
            for i in range(13, 45):
                ws[f'E{i}'].value = float(volumes_in_bn[i-13])

        ws['D9']=f'সর্বোচ্চ অনুমোদিত ত্রুটির পরিমাণঃ {str(allowable_limit)} {unit}'
        ws['J3']=f'পণ্যের নাম: {name}'
        ws['J4']=f'ব্রান্ড: {brand}'
        ws['J5']=f'মোড়কজাতকরণের তারিখ: {inspection_date}'
        ws['D5']=f'নামঃ {company_details["title"]}'
        ws['D6']=f'ঠিকানাঃ {company_details["address"]}'
        ws['K8']=str(size)+" "+unit
        ws['K9']="" if varient=="none" else varient
        
